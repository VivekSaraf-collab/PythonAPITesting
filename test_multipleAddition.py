import json
import pytest
import requests
import jsonpath
import openpyxl


def test_MultipleStudent():
    API_URL = "http://thetestingworldapi.com/api/studentsDetails"
    file = open('C:\\LocalRepos\\APIAutomationPython\\json\\End2End\\studentDetails.json', 'r')
    json_request = json.loads(file.read())

    wk = openpyxl.load_workbook("C:\\LocalRepos\\APIAutomationPython\\json\\End2End\\SD.xlsx")
    sh = wk('Sheet1')
    rows = sh.max_row

    for i in range(2, rows + 1):
        cell_firstName = sh.cell(row=i, column=1)
        cell_middleName = sh.cell(row=i, column=2)
        cell_lastName = sh.cell(row=i, column=3)
        cell_date_of_birth = sh.cell(row=i, column=4)
        json_request['firstName'] = cell_firstName.value
        json_request['middleName'] = cell_middleName.value
        json_request['lastName'] = cell_lastName.value
        json_request['date_of_birth'] = cell_date_of_birth.value

        response = requests.post(API_URL, json_request)

        print(response.text)
        print(response.status_code)
        assert response.status_code == 201
