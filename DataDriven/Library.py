import json
import jsonpath
import requests
import openpyxl

class common:

    def __init__(self,FileNamepath, SheetName):
        global wk
        global sh
        wk = openpyxl.load_workbook(FileNamepath)
        sh = wk(SheetName)

    def fetch_rowCount(self):
        rows = sh.max_row
        return rows

    def fetch_columnCount(self):
        col = sh.max_column
        return col

    def fetch_keyNames(self):
        c = sh.max_column
        li={}
        for i in range(1,c+1)
            cell = sh.cell(row = 1, column = i)
            li.insert(i-1,cell.value)

        return li

    def update_requestWithData(self,rowNumber, jsonRequest,keyList):
        c = sh.max_column
        for i in range(1, c+1):
            cell = sh.cell(row = rowNumber, column = i)
            jsonRequest[keyList[i-1]]=cell.value

        return jsonRequest